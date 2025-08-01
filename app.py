

import os
import io
from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import pandas as pd
import streamlit as st

# Load environment variables
load_dotenv()
AZURE_BLOB_CONNECTION_STRING = os.getenv("AZURE_BLOB_CONNECTION_STRING")
AZURE_BLOB_KEY = os.getenv("AZURE_BLOB_KEY")
AZURE_BLOB_CONTAINER = os.getenv("AZURE_BLOB_CONTAINER", "budget-docs")
AZURE_FORMRECOGNIZER_ENDPOINT = os.getenv("AZURE_FORMRECOGNIZER_ENDPOINT")
AZURE_FORMRECOGNIZER_KEY = os.getenv("AZURE_FORMRECOGNIZER_KEY")

def get_blob_service_client():
    if AZURE_BLOB_CONNECTION_STRING:
        return BlobServiceClient.from_connection_string(AZURE_BLOB_CONNECTION_STRING)
    elif AZURE_BLOB_KEY:
        st.warning("Connection string is preferred. Using key directly may not work for all setups.")
        return BlobServiceClient(account_url=f"https://{AZURE_BLOB_CONTAINER}.blob.core.windows.net", credential=AZURE_BLOB_KEY)
    else:
        st.error("No Azure Blob Storage credentials found in .env.")
        return None

def list_blobs_in_container(container_name):
    blob_service_client = get_blob_service_client()
    if not blob_service_client:
        return []
    try:
        container_client = blob_service_client.get_container_client(container_name)
        blobs = container_client.list_blobs()
        return [blob.name for blob in blobs if '/' not in blob.name and '\\' not in blob.name]
    except Exception as e:
        st.error(f"Error listing blobs: {e}")
        return []

def download_blob_to_bytes(blob_name, container_name):
    blob_service_client = get_blob_service_client()
    if not blob_service_client:
        return None
    try:
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
        return blob_client.download_blob().readall()
    except Exception as e:
        st.error(f"Error downloading blob '{blob_name}': {e}")
        return None

def analyze_with_document_intelligence(file_bytes, file_name, user_selected_model):
    if not (AZURE_FORMRECOGNIZER_ENDPOINT and AZURE_FORMRECOGNIZER_KEY):
        st.error("Azure Document Intelligence credentials not set in .env.")
        return None, None, None, None, None
    
    client = DocumentAnalysisClient(
        endpoint=AZURE_FORMRECOGNIZER_ENDPOINT,
        credential=AzureKeyCredential(AZURE_FORMRECOGNIZER_KEY)
    )
    
    # STEP 1: Layout Analysis
    layout_tables = []
    merged_table = None
    try:
        st.sidebar.info("🔍 STEP 1: Running Layout Analysis...")
        poller = client.begin_analyze_document("prebuilt-layout", file_bytes)
        layout_result = poller.result()
        st.sidebar.success("✅ STEP 1: Layout analysis completed")
        
        # Process all tables from all pages
        if hasattr(layout_result, "tables") and layout_result.tables:
            st.sidebar.info(f"📊 Found {len(layout_result.tables)} tables in layout...")
            for idx, table in enumerate(layout_result.tables):
                st.sidebar.info(f"📄 Processing table {idx+1}/{len(layout_result.tables)} (Page {table.bounding_regions[0].page_number if table.bounding_regions else 'Unknown'})")
                rows = []
                for row_idx in range(table.row_count):
                    row = []
                    for col_idx in range(table.column_count):
                        cell = next((c for c in table.cells if c.row_index == row_idx and c.column_index == col_idx), None)
                        row.append(cell.content if cell else "")
                    rows.append(row)
                df = pd.DataFrame(rows)
                layout_tables.append(df)
            
            # Merge related tables if possible
            if layout_tables:
                st.sidebar.info("🔗 Checking for table merging opportunities...")
                first_cols = [tuple(df.columns) for df in layout_tables]
                if all(cols == first_cols[0] for cols in first_cols):
                    merged_table = pd.concat(layout_tables, ignore_index=True)
                    st.sidebar.success("✅ Tables merged successfully")
                else:
                    merged_table = None
                    st.sidebar.info("ℹ️ Tables have different structures - keeping separate")
        else:
            st.sidebar.warning("⚠️ No tables found in layout analysis")
    except Exception as e:
        st.sidebar.error(f"❌ STEP 1 failed: {e}")
        st.error(f"Layout analysis failed: {e}")
        return None, None, None, None, None

    # STEP 2: Use User-Selected Model for Label Extraction
    model_id = user_selected_model
    st.sidebar.info(f"👤 STEP 2: Using user-selected model: {model_id}")
    
    # Run prebuilt model analysis
    model_labels = {}
    try:
        st.sidebar.info(f"🤖 STEP 2: Extracting labels using {model_id}...")
        poller = client.begin_analyze_document(model_id, file_bytes)
        result = poller.result()
        st.sidebar.success(f"✅ STEP 2: Label extraction completed")
        
        if hasattr(result, "documents") and result.documents:
            doc = result.documents[0]
            for k, v in doc.fields.items():
                if v and hasattr(v, "value") and v.value is not None:
                    model_labels[k] = v.value
                elif v is not None:
                    model_labels[k] = str(v)
            
            st.sidebar.success(f"✅ Found {len(model_labels)} non-null labels")
        else:
            st.sidebar.warning("⚠️ No documents found in prebuilt model results")
    except Exception as e:
        st.sidebar.error(f"❌ STEP 2 failed: {e}")
        st.error(f"Prebuilt model analysis failed: {e}")

    # STEP 3: Extract Total Budget from Labels
    st.sidebar.info("💰 STEP 3: Extracting total budget from labels...")
    
    budget_synonyms = [
        "total", "amount", "grand_total", "net_amount", "subtotal", "sub_total",
        "final_total", "overall_total", "total_amount", "total_cost", "total_price",
        "sum", "budget_amount", "budget_total", "cost", "price", "invoice_total",
        "bill_total", "payment", "due", "balance", "gross", "net", "final",
        "overall", "aggregate", "cumulative", "grand_sum", "total_tax", "amount_due"
    ]
    
    budget_data = []
    final_budget = None
    
    # Extract budget-related labels (ignore null values)
    for label, value in model_labels.items():
        if value is not None and str(value).strip() and str(value).lower() != 'none':
            label_lower = label.lower().replace(" ", "_")
            if any(synonym in label_lower for synonym in budget_synonyms):
                budget_data.append({
                    'Label': label,
                    'Value': value,
                    'Type': f'User Selected ({model_id})'
                })
                
                # Determine if this is likely the final budget
                priority_score = 0
                high_priority_terms = ["total", "grand_total", "amount_due", "final_total", "net_amount"]
                for i, term in enumerate(high_priority_terms):
                    if term in label_lower:
                        priority_score = len(high_priority_terms) - i
                        break
                
                if not final_budget or priority_score > final_budget.get('priority', 0):
                    final_budget = {
                        'label': label,
                        'value': value,
                        'priority': priority_score,
                        'source': f'User Selected ({model_id})'
                    }
    
    if budget_data:
        st.sidebar.success(f"✅ STEP 3: Found {len(budget_data)} budget-related entries")
    else:
        st.sidebar.warning("⚠️ STEP 3: No budget data found in labels")
    
    st.sidebar.success("🎉 All analysis steps completed!")
    
    return layout_tables, merged_table, model_labels, budget_data, final_budget

st.set_page_config(page_title="Budget Extractor", layout="centered")
st.title("Budget Extractor")

# File source selection
file_source = st.sidebar.radio("Choose file source:", ["Azure Blob Storage", "Local Upload"])

# Manual model selection
st.sidebar.subheader("🤖 Model Selection")
model_choice = st.sidebar.selectbox(
    "Select Azure Document Intelligence Model:",
    ["prebuilt-invoice", "prebuilt-document", "prebuilt-receipt"],
    index=0,  # Default to prebuilt-invoice
    help="Choose the model that best fits your document type"
)

selected_file = None
uploaded_file = None
file_bytes = None

if file_source == "Azure Blob Storage":
    st.sidebar.header("Azure Blob Storage")
    blob_files = list_blobs_in_container(AZURE_BLOB_CONTAINER)
    selected_file = st.sidebar.selectbox("Select a file to process:", blob_files)
else:
    st.sidebar.header("Local File Upload")
    uploaded_file = st.sidebar.file_uploader(
        "Choose a file to upload",
        type=['pdf', 'png', 'jpg', 'jpeg', 'bmp', 'tiff', 'xlsx', 'xls', 'csv', 'json']
    )

# Check if file is already processed to show appropriate button text
current_file = selected_file if file_source == "Azure Blob Storage" else (uploaded_file.name if uploaded_file else None)
cache_key = f"processed_{current_file}" if current_file else None
is_processed = cache_key and cache_key in st.session_state
button_text = "RE-RUN" if is_processed else "RUN"

# Only show the run button if a file is selected/uploaded
if current_file:
    run_analysis = st.sidebar.button(button_text)
else:
    run_analysis = False
    st.sidebar.info("Please select or upload a file to process.")

if current_file and run_analysis:
    st.info(f"Processing: {current_file}")
    
    # Clear previous analysis status
    st.sidebar.markdown("---")
    st.sidebar.subheader("📊 Analysis Progress")
    
    # Get file bytes based on source
    if file_source == "Azure Blob Storage":
        st.sidebar.info("📥 Downloading file from Azure Blob Storage...")
        file_bytes = download_blob_to_bytes(selected_file, AZURE_BLOB_CONTAINER)
        if not file_bytes:
            st.error("Could not download the selected file.")
            st.stop()
        file_name = selected_file
        st.sidebar.success("✅ File downloaded successfully")
    else:  # Local upload
        st.sidebar.info("📁 Reading uploaded file...")
        file_bytes = uploaded_file.read()
        file_name = uploaded_file.name
        st.sidebar.success("✅ File read successfully")
    
    # Show selected prebuilt model in sidebar during processing
    ext = os.path.splitext(file_name)[1].lower()
    if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"]:
        st.sidebar.info("🔍 Will analyze layout first, then select optimal prebuilt model")
    
    use_cache = False  # Always reprocess when button is clicked
    final_budget, layout_tables, model_budget, model_id = None, None, None, None
    excel_sheets, csv_df, json_data = None, None, None
    
    if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"]:
        st.sidebar.info("🔍 Starting 4-Step Document Intelligence Process...")
        layout_tables, merged_table, model_labels, budget_data, final_budget = analyze_with_document_intelligence(file_bytes, file_name, model_choice)
    elif ext in [".xlsx", ".xls"]:
        # STEP 4: Excel files - Skip analysis and show directly
        st.sidebar.info("📊 STEP 4: Excel file detected - Skipping analysis...")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            excel_sheets = {}
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                rows = [row for row in ws.iter_rows(values_only=True)]
                df = pd.DataFrame(rows)
                excel_sheets[sheetname] = df
            st.sidebar.success(f"✅ STEP 4: Excel file processed - {len(excel_sheets)} sheet(s) found")
        except Exception as e:
            st.sidebar.error(f"❌ Excel processing failed: {e}")
            st.error(f"Error reading Excel file: {e}")
    elif ext == ".csv":
        st.sidebar.info("📄 Processing CSV file...")
        try:
            import io
            import csv
            reader = csv.reader(io.StringIO(file_bytes.decode('utf-8')))
            rows = [row for row in reader]
            csv_df = pd.DataFrame(rows)
            st.sidebar.success(f"✅ CSV file processed - {len(csv_df)} rows found")
        except Exception as e:
            st.sidebar.error(f"❌ CSV processing failed: {e}")
            st.error(f"Error reading CSV file: {e}")
    elif ext == ".json":
        st.sidebar.info("🔧 Processing JSON file...")
        try:
            import json
            json_data = json.loads(file_bytes.decode('utf-8'))
            st.sidebar.success("✅ JSON file processed successfully")
        except Exception as e:
            st.sidebar.error(f"❌ JSON processing failed: {e}")
            st.error(f"Error reading JSON file: {e}")
    
    # Final status
    st.sidebar.markdown("---")
    st.sidebar.success("🎉 All processing completed!")
    
    # Update cache with new results
    st.session_state[cache_key] = {
        'ext': ext,
        'final_budget': final_budget if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"] else None,
        'layout_tables': layout_tables if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"] else None,
        'model_labels': model_labels if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"] else None,
        'budget_data': budget_data if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"] else None,
        'merged_table': merged_table if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"] else None,
        'excel_sheets': excel_sheets,
        'csv_df': csv_df,
        'json_data': json_data
    }

    # Display results from cache or fresh
    if ext in [".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff"]:
        # Results from 4-Step Process
        st.subheader("🎯 Final Budget Amount")
        if final_budget:
            st.success(f"**{final_budget['label']}**: {final_budget['value']}")
            st.info(f"Source: {final_budget['source']} | Priority Score: {final_budget['priority']}")
        else:
            st.warning("No final budget amount identified.")
        
        # STEP 3 Results: Budget Data Table
        st.subheader("💰 Budget Data Analysis (Step 3)")
        if budget_data:
            budget_df = pd.DataFrame(budget_data)
            st.write(f"**Found {len(budget_df)} budget-related entries (excluding null values):**")
            st.dataframe(budget_df, use_container_width=True, hide_index=True)
            
            # Download budget data
            csv_data = budget_df.to_csv(index=False)
            st.download_button(
                label="Download Budget Data as CSV",
                data=csv_data,
                file_name=f"{current_file}_budget_analysis.csv",
                mime="text/csv"
            )
            
            json_data = budget_df.to_json(orient='records', force_ascii=False, indent=2)
            st.download_button(
                label="Download Budget Data as JSON",
                data=json_data,
                file_name=f"{current_file}_budget_analysis.json",
                mime="application/json"
            )
        else:
            st.info("No budget-related data found in labels.")
        
        # STEP 2 Results: All Labels (Optional view)
        if model_labels:
            with st.expander(f"🏷️ All Extracted Labels (Step 2) - {len(model_labels)} labels"):
                labels_df = pd.DataFrame([
                    {'Label': k, 'Value': v} 
                    for k, v in model_labels.items()
                ])
                st.dataframe(labels_df, use_container_width=True, hide_index=True)
        
        # STEP 1 Results: Layout Tables (Optional view)
        if layout_tables:
            with st.expander(f"📊 Layout Analysis Results (Step 1) - {len(layout_tables)} tables"):
                if merged_table is not None:
                    st.write("**Merged Table (All Pages):**")
                    st.dataframe(merged_table, use_container_width=True, hide_index=True)
                else:
                    for idx, df in enumerate(layout_tables):
                        st.write(f"**Table {idx+1}:**")
                        st.dataframe(df, use_container_width=True, hide_index=True)
    
    elif ext in [".xlsx", ".xls"]:
        # STEP 4: Excel Direct Display
        st.subheader("📊 Excel File - Direct Display (Step 4)")
        if excel_sheets:
            for sheetname, df in excel_sheets.items():
                st.write(f"**Sheet: {sheetname}** ({len(df)} rows, {len(df.columns)} columns)")
                st.dataframe(df, use_container_width=True)
                
                # JSON download option for Excel
                json_data = df.to_json(orient='records', force_ascii=False, indent=2)
                st.download_button(
                    label=f"Download {sheetname} as JSON",
                    data=json_data,
                    file_name=f"{current_file}_{sheetname}.json",
                    mime="application/json"
                )
                
                # Also provide CSV option
                csv_data = df.to_csv(index=False)
                st.download_button(
                    label=f"Download {sheetname} as CSV",
                    data=csv_data,
                    file_name=f"{current_file}_{sheetname}.csv",
                    mime="text/csv"
                )
        else:
            st.write("No data found in Excel file.")
    elif ext == ".csv":
        st.subheader("CSV file detected. Showing complete data:")
        if csv_df is not None:
            st.write(f"**Complete CSV Data** ({len(csv_df)} rows)")
            # Display all rows without limit
            st.dataframe(csv_df, use_container_width=True, hide_index=True)
            
            # Download button for processed CSV
            csv_data = csv_df.to_csv(index=False)
            st.download_button(
                label="Download Processed CSV",
                data=csv_data,
                file_name=f"{current_file}_processed.csv",
                mime="text/csv"
            )
        else:
            st.write("No data found in CSV file.")
    elif ext == ".json":
        st.subheader("JSON file detected. Showing complete data:")
        if json_data is not None:
            st.json(json_data)
            
            # Download button for processed JSON
            json_str = json.dumps(json_data, indent=2, ensure_ascii=False)
            st.download_button(
                label="Download Processed JSON",
                data=json_str,
                file_name=f"{current_file}_processed.json",
                mime="application/json"
            )
        else:
            st.write("No data found in JSON file.")
    else:
        st.warning(f"Unsupported file type: {ext}")
